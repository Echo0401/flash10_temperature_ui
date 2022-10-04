# -*- coding: utf-8 -*-
import os
import sys
import time
import scipy.signal
import numpy as np
from openpyxl import load_workbook
from openpyxl.drawing.image import Image

def get_project_path():
    """
    获取项目的根目录
    :return: 根目录
    """
    # 判断调试模式
    debug_vars = dict((a, b) for a, b in os.environ.items() if a.find('IPYTHONENABLE') >= 0)
    # 根据不同场景获取根目录
    if len(debug_vars) > 0:
        """当前为debug运行时"""
        project_path = sys.path[2]
    elif getattr(sys, 'frozen', False):
        """当前为exe运行时"""
        project_path = os.getcwd()
    else:
        """正常执行"""
        project_path = sys.path[1]
    # 替换斜杠
    project_path = project_path.replace("\\", "/")
    return project_path


ProjectPath = get_project_path()

def get_file_absolute_path(fileName):
    """
    根据文件名获取资源文件路径
    """
    return ProjectPath + fileName

qc_result_template_file_path = get_file_absolute_path('/data/qc_result_template.xlsx')

class QualityReport(object):

    # --------------------------------温度测试相关-----------------------------------------------
    @staticmethod
    def min_d(s):
        # 传入可迭代对象为空时，必须指定参数default，用来返回默认值
        return min(s, default=0)

    @staticmethod
    def max_d(s):
        # 传入可迭代对象为空时，必须指定参数default，用来返回默认值
        return max(s, default=1000)

    @staticmethod
    def pf(s: bool):
        return "PASS" if s else "FAIL"

    @staticmethod
    def ramp(experiment_data):
        """
        平均升温速率
        平均降温速率
        修改升降温速率的算法，使时间完全落在 [49.5, 50.5] - [89.5, 90.5]的区间中
        """
        # 根据数据库更新了ramp()的读取
        ramp_dic = {}
        ramp_data = {}
        for data in experiment_data:
            ramp_data[data["position"]] = data["temperature_data"]

        ramp_data = {key: np.interp(np.arange(0, len(value) - 1, 0.01), range(len(value)), value) for key, value in
                     ramp_data.items()}

        for position, r_data in ramp_data.items():
            up_start, up_end, down_start, down_end = 0, 0, 0, 0
            step = 0
            for i, value in enumerate(r_data):
                if step == 0:
                    if value <= 47:
                        step = 1
                elif step == 1:
                    if value >= 49.5:
                        up_start = i
                        step = 2
                elif step == 2:
                    if value >= 90:
                        up_end = i
                        step = 3
                elif step == 3:
                    if value >= 95:
                        step = 4
                elif step == 4:
                    if value <= 90.5:
                        down_start = i
                        step = 5
                elif step == 5:
                    if value <= 50:
                        down_end = i
                        break

            up_speed = (r_data[up_end] - r_data[up_start]) / ((up_end - up_start) * 0.001)
            down_speed = (r_data[down_start] - r_data[down_end]) / ((down_end - down_start) * 0.001)
            ramp_dic[position] = [round(up_speed, 2), round(down_speed, 2)]

        return ramp_dic

    @staticmethod
    def time_accuracy(experiment_data):
        """
        温度持续时间准确度
        """
        time_accuracy_info = {}  # key 位置
        for data in experiment_data:
            temperature_data = data["temperature_data"]
            position = data["position"]
            time_interval = data["time_interval"]
            end_time_list_temp = []
            durations = list()
            temperatures = [95] * 5
            value_index = 0
            for round_index, temperature in enumerate(temperatures):
                found_start = False
                start_time = 0
                while True:
                    if not found_start:
                        if temperature_data[value_index] > temperature - 0.5 and \
                                temperature_data[value_index + 20] > temperature - 0.5:
                            # 过滤波峰
                            # start_time = (value_index - 5) * 0.1
                            # peak_95 = np.array(r_data[value_index:value_index + 1200])
                            # index_start_95, _ = scipy.signal.find_peaks(peak_95, distance=1200)
                            # start_95 = index_start_95[0] + value_index + 19
                            # start_time = start_95- 5
                            start_time = value_index - 5
                            found_start = True
                            value_index += 10
                        else:
                            value_index += 10
                    else:
                        if temperature_data[value_index] < temperature - 0.5 and \
                                temperature_data[value_index + 20] < temperature - 0.5:
                            # end_time = value_index * 0.1
                            end_time = value_index
                            end_time_list_temp.append(end_time + 20)
                            durations.append((start_time + 2, end_time + 2))
                            break
                        value_index += 1

            try:
                avg_time = sum([abs(e - s) for (e, s) in durations]) / len(durations)
            except ZeroDivisionError:
                avg_time = 0

            time_accuracy_info[position] = {
                "end_time": max(end_time_list_temp),
                "time_accuracy": abs(round((avg_time * time_interval - 120) / 120, 4)) * 100}

        return time_accuracy_info

    @staticmethod
    def accuracy(temperature_data):
        """
        模块控温精度
        """
        time_accuracy_info = QualityReport.time_accuracy(temperature_data)
        control_result = {}
        for data in temperature_data:
            position = data["position"]
            r_data_tmp = data["temperature_data"]
            r_data = r_data_tmp[0: len(r_data_tmp) - 100]
            control_result[position] = {}
            try:
                accuracy_data = {55: list(), 72: list(), 95: list()}
                step = 0
                start_55, start_72, start_95 = 0, 0, 0
                for i, value in enumerate(r_data):
                    if i > time_accuracy_info[position]["end_time"]:
                        if step == 0:
                            if value < 60:
                                step = 1
                        elif step == 1:
                            if 54.5 < value <= 55.5:
                                # 过滤波谷
                                trough_55 = np.array(r_data[i:i + 1200])
                                index_start_55, _ = scipy.signal.find_peaks(-trough_55, distance=1200)
                                start_55 = index_start_55[0] + i + 19
                                if start_55 - i > 100:
                                    start_55 = i + 40
                                step = 2
                        elif step == 2:
                            if 71.5 < value <= 72.5:
                                # 过滤波峰
                                peak_72 = np.array(r_data[i:i + 1200])
                                index_start_72, _ = scipy.signal.find_peaks(peak_72, distance=1200)
                                start_72 = index_start_72[0] + i + 19
                                if start_72 - i > 100:
                                    start_72 = i + 40
                                step = 3
                        elif step == 3:
                            if 94.5 < value <= 95.5:
                                # 过滤波峰
                                peak_95 = np.array(r_data[i:i + 1200])
                                index_start_95, _ = scipy.signal.find_peaks(peak_95, distance=1200)
                                start_95 = index_start_95[0] + i + 19
                                if start_95 - i > 100:
                                    start_95 = i + 40
                                step = 4
                        elif step == 4:
                            accuracy_data[55].append(start_55)
                            accuracy_data[72].append(start_72)
                            accuracy_data[95].append(start_95)
                            if len(accuracy_data[55]) >= 5:
                                break
                            else:
                                step = 0

                # 计算模块控温精度
                control_result[position]["control_accuracy_55"] = abs(
                    QualityReport.max_d([(QualityReport.max_d(r_data[start_time + 100:start_time + 400]) - QualityReport.min_d(
                        r_data[start_time + 100:start_time + 400])) / 2 for start_time in accuracy_data[55]]))
                control_result[position]["control_accuracy_72"] = abs(
                    QualityReport.max_d([(QualityReport.max_d(r_data[start_time + 100:start_time + 400]) - QualityReport.min_d(
                        r_data[start_time + 100:start_time + 400])) / 2 for start_time in accuracy_data[72]]))
                control_result[position]["control_accuracy_95"] = abs(
                    QualityReport.max_d([(QualityReport.max_d(r_data[start_time + 100:start_time + 400]) - QualityReport.min_d(
                        r_data[start_time + 100:start_time + 400])) / 2 for start_time in accuracy_data[95]]))

                # 计算温度准确度
                start_time_55 = accuracy_data[55][0]
                start_time_72 = accuracy_data[72][0]
                start_time_95 = accuracy_data[95][0]
                control_result[position]["temperature_accuracy_55"] = abs(np.average(
                    [r_data[start_time_55 + 100], r_data[start_time_55 + 200], r_data[start_time_55 + 300],
                     r_data[start_time_55 + 400], r_data[start_time_55 + 500],
                     r_data[start_time_55 + 600]]) - 55)
                control_result[position]["temperature_accuracy_72"] = abs(np.average(
                    [r_data[start_time_72 + 100], r_data[start_time_72 + 200], r_data[start_time_72 + 300],
                     r_data[start_time_72 + 400], r_data[start_time_72 + 500],
                     r_data[start_time_72 + 600]]) - 55)
                control_result[position]["temperature_accuracy_95"] = abs(np.average(
                    [r_data[start_time_95 + 100], r_data[start_time_95 + 200], r_data[start_time_95 + 300],
                     r_data[start_time_95 + 400], r_data[start_time_95 + 500],
                     r_data[start_time_95 + 600]]) - 55)

                # 计算模块温度均匀性
                control_result[position]["temperature_accuracy_55"] = r_data[start_time_55 + 600]
                control_result[position]["temperature_accuracy_72"] = r_data[start_time_72 + 600]
                control_result[position]["temperature_accuracy_95"] = r_data[start_time_95 + 600]
            except Exception:
                raise
        uniformity_55, uniformity_72, uniformity_95 = [], [], []
        for position, info in control_result.items():
            uniformity_55.append(info["temperature_accuracy_55"])
            uniformity_72.append(info["temperature_accuracy_72"])
            uniformity_95.append(info["temperature_accuracy_95"])

        control_result["uniformity_55"] = QualityReport.max_d(uniformity_55) - QualityReport.min_d(uniformity_55)
        control_result["uniformity_72"] = QualityReport.max_d(uniformity_72) - QualityReport.min_d(uniformity_72)
        control_result["uniformity_95"] = QualityReport.max_d(uniformity_95) - QualityReport.min_d(uniformity_95)
        return control_result

    @classmethod
    def export_temperature_report(cls, experiment_data, export_path):
        """
        生成温度的报告并导出
        """
        # try:
        if True:
            ramp_dic = QualityReport.ramp(experiment_data)
            time_accuracy_info = QualityReport.time_accuracy(experiment_data)
            accuracy_info = QualityReport.accuracy(experiment_data)
            wb = load_workbook(qc_result_template_file_path)
            sheet_name_list = ['温度测试']
            ws = wb['温度测试']


            ws.cell(2, 1, value=f"仪器SN号：")
            ws.cell(3, 1, value=f"PT100温度工装：")
            ws.cell(6, 3, value=round(accuracy_info["uniformity_55"], 2))
            ws.cell(6, 4, value=QualityReport.pf(accuracy_info["uniformity_55"] <= 1))
            ws.cell(7, 3, value=round(accuracy_info["uniformity_72"], 2))
            ws.cell(7, 4, value=QualityReport.pf(accuracy_info["uniformity_72"] <= 1))
            ws.cell(8, 3, value=round(accuracy_info["uniformity_95"], 2))
            ws.cell(8, 4, value=QualityReport.pf(accuracy_info["uniformity_95"] <= 1))

            for module in range(0, 4):
                ws.cell(12, ((module + 1) * 2 + 1), value=round(ramp_dic[module][0], 4))
                ws.cell(12, ((module + 2) * 2), value=QualityReport.pf(ramp_dic[module][0] >= 15))
                ws.cell(13, ((module + 1) * 2 + 1), value=round(ramp_dic[module][1], 4))
                ws.cell(13, ((module + 2) * 2), value=QualityReport.pf(ramp_dic[module][1] >= 15))

            for module in range(0, 4):
                ws.cell(14, ((module + 1) * 2 + 1), value=round(accuracy_info[module]["control_accuracy_55"], 2))
                ws.cell(14, ((module + 2) * 2), value=QualityReport.pf(accuracy_info[module]["control_accuracy_55"] <= 0.5))

                ws.cell(15, ((module + 1) * 2 + 1), value=round(accuracy_info[module]["control_accuracy_72"], 2))
                ws.cell(15, ((module + 2) * 2), value=QualityReport.pf(accuracy_info[module]["control_accuracy_72"] <= 0.5))

                ws.cell(16, ((module + 1) * 2 + 1), value=round(accuracy_info[module]["control_accuracy_95"], 2))
                ws.cell(16, ((module + 2) * 2), value=QualityReport.pf(accuracy_info[module]["control_accuracy_95"] <= 0.5))

            for module in range(0, 4):
                ws.cell(17, ((module + 1) * 2 + 1),
                        value=abs(accuracy_info[module]["temperature_accuracy_55"] - 55))
                ws.cell(17, ((module + 2) * 2),
                        value=QualityReport.pf(abs(accuracy_info[module]["temperature_accuracy_55"] - 55) <= 0.5))
                ws.cell(18, ((module + 1) * 2 + 1),
                        value=abs(accuracy_info[module]["temperature_accuracy_72"] - 72))
                ws.cell(18, ((module + 2) * 2),
                        value=QualityReport.pf(abs(accuracy_info[module]["temperature_accuracy_72"] - 72) <= 0.5))
                ws.cell(19, ((module + 1) * 2 + 1),
                        value=abs(accuracy_info[module]["temperature_accuracy_95"] - 95))
                ws.cell(19, ((module + 2) * 2),
                        value=QualityReport.pf(abs(accuracy_info[module]["temperature_accuracy_95"] - 95) <= 0.5))

            for module in range(0, 4):

                ws.cell(20, ((module + 1) * 2 + 1), value=f'{round(time_accuracy_info[module]["time_accuracy"], 2)}%')
                ws.cell(20, ((module + 2) * 2),
                        value=QualityReport.pf(round(time_accuracy_info[module]["time_accuracy"], 2) <= 5))

            serial_list = []
            curve_list = []

            for data in experiment_data:
                curve_list.append(data["temperature_data"])
                serial_list.append(data["position"])


            sheet_names = wb.get_sheet_names()
            for item in sheet_names:
                if item not in sheet_name_list:
                    remove_sheet = wb.get_sheet_by_name(item)
                    wb.remove_sheet(remove_sheet)

            wb.save(export_path)

            return True, "报告生成成功"
        # except Exception as err:
        #     return False, "报告生成失败"

if __name__ == '__main__':
    qualityreport = QualityReport()
    # experiment_data =
    # export_path =
    qualityreport.export_temperature_report()

