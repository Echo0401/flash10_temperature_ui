# -*- coding: utf-8 -*-

import multiprocessing
import sys
from datetime import datetime
from PySide2 import QtWidgets, QtCore
from openpyxl import load_workbook
from PyCRC.CRC32 import CRC32
import numpy as np
import scipy.signal
import xlrd


crc = CRC32()

"""
从生成的Excel文档中读取温度数据，并生成Excel文档
"""

qc_result_template_file_path = f"data/qc_result_template.xlsx"

class MainWindow(QtWidgets.QMainWindow):
    operation_signal = QtCore.Signal(bool, dict)

    def __init__(self):
        super(MainWindow, self).__init__()
        version = "V1.0.3"

        box_main = QtWidgets.QVBoxLayout()
        central_widget = QtWidgets.QWidget()
        central_widget.setLayout(box_main)
        self.setCentralWidget(central_widget)

        self.setWindowTitle(f"温度数据处理-{version}")
        self.group_max_min = QtWidgets.QGroupBox("生成温度报告")
        self.group_max_min.setFixedWidth(500)

        box_group = QtWidgets.QHBoxLayout()
        box_group.addWidget(self.group_max_min)
        box_main.addLayout(box_group)

        # ------------生成温度报告--------------
        v_max_min = QtWidgets.QVBoxLayout()

        f_max_min = QtWidgets.QFormLayout()
        self.btn_select_file = QtWidgets.QPushButton("选择文件")
        self.btn_select_file.clicked.connect(self.select_file)
        self.label_select_file = QtWidgets.QLabel()
        self.label_select_file.setMinimumHeight(60)
        self.label_select_file.setWordWrap(True)

        self.btn_export_file = QtWidgets.QPushButton("导出路径")
        self.btn_export_file.clicked.connect(self.export_file)
        self.label_export_file = QtWidgets.QLabel()
        self.label_export_file.setMinimumHeight(60)
        self.label_export_file.setWordWrap(True)

        self.btn_get_max_min = QtWidgets.QPushButton("生成报告")
        self.btn_get_max_min.clicked.connect(self.temperature_report)
        f_max_min.addRow(QtWidgets.QLabel("原始数据  "), self.btn_select_file)
        f_max_min.addRow(self.label_select_file)
        f_max_min.addRow(QtWidgets.QLabel("温度报告  "), self.btn_export_file)
        f_max_min.addRow(self.label_export_file)
        f_max_min.addRow(self.btn_get_max_min)
        v_max_min.addItem(f_max_min)
        v_max_min.addStretch()

        self.group_max_min.setLayout(v_max_min)


    def select_file(self):
        select_file, _ = QtWidgets.QFileDialog().getOpenFileName(self, self.tr("选择文件"), "./", "hex Files (*.xlsx);")
        if not select_file:
            self.label_select_file.setText("")
            return

        self.label_select_file.setText(select_file)

    def export_file(self):
        file_type = 'xlsx'
        default_file_name = 'temperature_' + datetime.now().strftime('%Y%m%d_%H%M%S')
        export_file, _ = QtWidgets.QFileDialog().getSaveFileName(self, self.tr("数据保存"),
                                                                 f"./{default_file_name}",
                                                                 f"{file_type} Files (*.{file_type}); ")

        if not export_file:
            self.label_export_file.setText("")
            return

        self.label_export_file.setText(export_file)

    def get_excel_data(self, filename):
        """
        获取温度循环的原始数据工厂版5个循环45-95；55-72-95
        """
        try:
            data = xlrd.open_workbook(filename)

            table = data.sheet_by_index(0)
            colAmount = table.ncols
            select_data_row = []
            for colIndex in range(colAmount):
                select_data_row.append(table.col_values(colIndex))
            for i, data in enumerate(select_data_row):
                if i > 0:
                    select_data_row[i] = data[1:]
            select_data_row = select_data_row[1:]
            select_data = []
            for temperature_list_row in select_data_row:
                temperature_list = []
                for temperature_row in temperature_list_row:
                    if type(temperature_row) == float:
                        temperature_list.append(temperature_row)
                select_data.append(temperature_list)
        except Exception:
            raise
        return select_data

    def temperature_report(self):
        select_path = self.label_select_file.text()
        export_path = self.label_export_file.text()
        select_data = self.get_excel_data(select_path)
        self._export_file_temperature(select_data=select_data, export_path=export_path)

    @staticmethod
    def max_d(s):
        # 传入可迭代对象为空时，必须指定参数default，用来返回默认值
        return max(s, default=1000)

    @staticmethod
    def min_d(s):
        # 传入可迭代对象为空时，必须指定参数default，用来返回默认值
        return min(s, default=0)

    @staticmethod
    def pf(s: bool):
        return "PASS" if s else "FAIL"


    @staticmethod
    def ramp(select_data):
        """
        平均升温速率
        平均降温速率
        """
        # 根据数据库更新了ramp()的读取
        ramp_dic = {}
        ramp_data = {}
        for i, data in enumerate(select_data):
            ramp_data[i] = data[1:]

        ramp_data = {key: np.interp(np.arange(0, len(value) - 1, 0.01), range(len(value)), value) for key, value in
                     ramp_data.items()}

        # for module in range(1, 5):
        for position, r_data in ramp_data.items():
            up_start, up_end, down_start, down_end = 0, 0, 0, 0
            step = 0
            for i, value in enumerate(r_data):
                if step == 0:
                    if value <= 47:
                        step = 1
                elif step == 1:
                    if value >= 50.5:
                        up_start = i
                        step = 2
                elif step == 2:
                    if value >= 90:
                        up_end = i
                        step = 3
                elif step == 3:
                    if value >= 94.5:
                        step = 4
                elif step == 4:
                    if value <= 89.5:
                        down_start = i
                        step = 5
                elif step == 5:
                    if value <= 50.5:
                        down_end = i
                        break


            print(f"————————————————————ramp____________________")
            print(f"行号 up_start + 2： {up_start + 2}")
            print(f"行号 up_end + 2： {up_end + 2}")
            print(f"行号 down_start + 2： {down_start + 2}")
            print(f"行号 down_end + 2： {down_end + 2}")
            up_speed = (r_data[up_end] - r_data[up_start]) / ((up_end - up_start) * 0.001)
            print(f"r_data[up_end]: {r_data[up_end]}, r_data[up_start]: {r_data[up_start]}, up_end : {up_end }, up_start: {up_start}")
            down_speed = (r_data[down_start] - r_data[down_end]) / ((down_end - down_start) * 0.001)
            print(
                f"r_data[down_start]: {r_data[down_start]}, r_data[down_end]: {r_data[down_end]}, down_end : {down_end}, down_start: {down_start}")
            ramp_dic[position] = [round(up_speed, 2), round(down_speed, 2)]
            # 行号 时间 温度
            # 行号 up_start + 2
            # 时间 up_start * 0.1
            # 温度 r_data[up_start]

        return ramp_dic

    @staticmethod
    def time_accuracy(select_data):
        """
        温度持续时间准确度
        """
        time_accuracy_info = {}  # 位置
        print(f"____________________select_data: ")
        for i, value in enumerate(select_data):
            print(value[:10])
        for position, temperature_data in enumerate(select_data[1:]):
            print(f"positon:  {position}")
            end_time_list_temp = []
            durations = list()
            temperatures = [95] * 5
            value_index = 0
            for round_index, temperature in enumerate(temperatures):
                found_start = False
                start_time = 0
                while True:
                    # module1,module2没进入循环
                    if not found_start:
                        # module1,module2没进入循环
                        if temperature_data[value_index] > temperature - 0.5 and \
                                temperature_data[value_index + 20] > temperature - 0.5:
                            # 过滤波峰
                            peak_95 = np.array(temperature_data[value_index:value_index + 1200])
                            index_start_95, _ = scipy.signal.find_peaks(peak_95, distance=1200)
                            start_time = index_start_95[0] + value_index + 19
                            if start_time - value_index > 100:
                                start_time = value_index + 40
                            found_start = True
                            value_index +=10
                        else:
                            value_index += 10
                    else:
                        if temperature_data[value_index] < temperature - 0.5 and \
                                temperature_data[value_index + 20] < temperature - 0.5:
                            # end_time = value_index * 0.1
                            # print(f"升温结束时间：{value_index}， 温度： {temperature_data[value_index]}")
                            end_time = value_index
                            end_time_list_temp.append(end_time + 20)
                            durations.append((start_time, end_time))
                            print(f"温度持续时间准确度：start_time + 22 行号：{start_time + 22}")
                            print(f"温度持续时间准确度：end_time + 22 行号：{end_time + 22}")
                            break
                        value_index += 1
            try:
                avg_time = sum([(e - s) for (e, s) in durations]) / len(durations)
            except ZeroDivisionError:
                avg_time = 0
            time_accuracy_info[position] = {
                "end_time": max(end_time_list_temp),
                "time_accuracy": round((avg_time * 0.1 - 120) / 120, 4) }
            print(f"模块控温精度：  {time_accuracy_info[position]}")

        return time_accuracy_info


    @staticmethod
    def accuracy(select_data):
        """
        模块控温精度
        """
        time_accuracy_info = MainWindow.time_accuracy(select_data[1:])
        print(f"time_accuracy_info: {time_accuracy_info}")
        control_result = {}
        for position, r_data_tmp in enumerate(select_data):
            print(f"控温精度：position: {position}")
            r_data = r_data_tmp[0: len(r_data_tmp)-100]
            control_result[position] = {}
            # try:
            accuracy_data = {55: list(), 72: list(), 95: list()}
            step = 0
            start_55, start_72, start_95 = 0, 0, 0
            print(f"________________报错原因")

            print(time_accuracy_info[position])
            print(time_accuracy_info[position]["end_time"])
            for i, value in enumerate(r_data):
                if i > time_accuracy_info[position]["end_time"]:
                    if step == 0:
                        if value < 60:
                            step = 1
                    elif step == 1:
                        if 54.5 < value <= 55.5:
                            # 过滤波谷
                            trough_55 = np.array(r_data[i:i + 1200])
                            index_start_55, _ = scipy.signal.find_peaks(-trough_55, distance=1200)
                            start_55 = index_start_55[0] + i + 19
                            if start_55 - i > 100:
                                start_55 = i + 40
                            step = 2
                    elif step == 2:
                        if 71.5 < value <= 72.5:
                            # 过滤波峰
                            peak_72 = np.array(r_data[i:i + 1200])
                            index_start_72, _ = scipy.signal.find_peaks(peak_72, distance=1200)
                            start_72 = index_start_72[0] + i + 19
                            if start_72 - i > 100:
                                start_72 = i + 40
                            step = 3
                    elif step == 3:
                        if 94.5 < value <= 95.5:
                            # 过滤波峰
                            peak_95 = np.array(r_data[i:i + 1200])
                            index_start_95, _ = scipy.signal.find_peaks(peak_95, distance=1200)
                            start_95 = index_start_95[0] + i + 19
                            if start_95 - i > 100:
                                start_95 = i + 40
                            step = 4
                    elif step == 4:

                        accuracy_data[55].append(start_55)
                        accuracy_data[72].append(start_72)
                        accuracy_data[95].append(start_95)
                        print(f"控温精度： start_55 + 2:{start_55 + 2}")
                        print(f"控温精度： start_72 + 2:{start_72 + 2}")
                        print(f"控温精度： start_95 + 2:{start_95 + 2}")

                        if len(accuracy_data[55]) >= 5:
                            break
                        else:
                            step = 0
            # 计算模块控温精度

            control_result[position]["control_accuracy_55"] = abs(
                MainWindow.max_d(
                    [(MainWindow.max_d(r_data[start_time + 100:start_time + 400]) - MainWindow.min_d(
                        r_data[start_time + 100:start_time + 400])) / 2 for start_time in accuracy_data[55]]))
            control_result[position]["control_accuracy_72"] = abs(
                MainWindow.max_d(
                    [(MainWindow.max_d(r_data[start_time + 100:start_time + 400]) - MainWindow.min_d(
                        r_data[start_time + 100:start_time + 400])) / 2 for start_time in accuracy_data[72]]))
            control_result[position]["control_accuracy_95"] = abs(
                MainWindow.max_d(
                    [(MainWindow.max_d(r_data[start_time + 100:start_time + 400]) - MainWindow.min_d(
                        r_data[start_time + 100:start_time + 400])) / 2 for start_time in accuracy_data[95]]))

            # 计算温度准确度
            start_time_55 = accuracy_data[55][0]
            start_time_72 = accuracy_data[72][0]
            start_time_95 = accuracy_data[95][0]
            control_result[position]["temperature_accuracy_55"] = abs(np.average(
                [r_data[start_time_55 + 100], r_data[start_time_55 + 200], r_data[start_time_55 + 300],
                 r_data[start_time_55 + 400], r_data[start_time_55 + 500], r_data[start_time_55 + 600]]) - 55)
            control_result[position]["temperature_accuracy_72"] = abs(np.average(
                [r_data[start_time_72 + 100], r_data[start_time_72 + 200], r_data[start_time_72 + 300],
                 r_data[start_time_72 + 400], r_data[start_time_72 + 500], r_data[start_time_72 + 600]]) - 72)
            control_result[position]["temperature_accuracy_95"] = abs(np.average(
                [r_data[start_time_95 + 100], r_data[start_time_95 + 200], r_data[start_time_95 + 300],
                 r_data[start_time_95 + 400], r_data[start_time_95 + 500], r_data[start_time_95 + 600]]) - 95)

            # 计算模块温度均匀性
            control_result[position]["single_module_uniformity_55"] = r_data[start_time_55 + 600]
            control_result[position]["single_module_uniformity_72"] = r_data[start_time_72 + 600]
            control_result[position]["single_module_uniformity_95"] = r_data[start_time_95 + 600]

            # except Exception:
            #     raise
        # 模块温度均匀性的计算
        uniformity_55, uniformity_72, uniformity_95 = [], [], []
        for position, info in control_result.items():
            uniformity_55.append(control_result[position]["single_module_uniformity_55"])
            uniformity_72.append(control_result[position]["single_module_uniformity_72"])
            uniformity_95.append(control_result[position]["single_module_uniformity_95"])

        control_result["uniformity_55"] = MainWindow.max_d(uniformity_55) - MainWindow.min_d(uniformity_55)
        control_result["uniformity_72"] = MainWindow.max_d(uniformity_72) - MainWindow.min_d(uniformity_72)
        control_result["uniformity_95"] = MainWindow.max_d(uniformity_95) - MainWindow.min_d(uniformity_95)
        return control_result


    def _export_file_temperature(self, select_data, export_path):

        ramp_dic = MainWindow.ramp(select_data)
        time_accuracy_info = MainWindow.time_accuracy(select_data)
        accuracy_info = MainWindow.accuracy(select_data)

        wb = load_workbook(qc_result_template_file_path)
        sheet_name_list = ['温度测试']
        ws = wb['温度测试']

        ws.cell(2, 1, value=f"仪器SN号：")
        ws.cell(3, 1, value=f"PT100温度工装：")
        ws.cell(4, 1, value=f"电脑编号：")
        ws.cell(7, 3, value=round(accuracy_info["uniformity_55"], 2))
        ws.cell(7, 4, value=MainWindow.pf(accuracy_info["uniformity_55"] <= 1))
        ws.cell(8, 3, value=round(accuracy_info["uniformity_72"], 2))
        ws.cell(8, 4, value=MainWindow.pf(accuracy_info["uniformity_72"] <= 1))
        ws.cell(9, 3, value=round(accuracy_info["uniformity_95"], 2))
        ws.cell(9, 4, value=MainWindow.pf(accuracy_info["uniformity_95"] <= 1))

        for module in range(0, 4):
            ws.cell(13, ((module + 1) * 2 + 1), value=round(ramp_dic[module][0], 4))
            ws.cell(13, ((module + 2) * 2), value=MainWindow.pf(ramp_dic[module][0] >= 15))
            ws.cell(14, ((module + 1) * 2 + 1), value=round(ramp_dic[module][1], 4))
            ws.cell(14, ((module + 2) * 2), value=MainWindow.pf(ramp_dic[module][1] >= 15))

        for module in range(0, 4):
            ws.cell(15, ((module + 1) * 2 + 1), value=round(accuracy_info[module]["control_accuracy_55"], 2))
            ws.cell(15, ((module + 2) * 2), value=MainWindow.pf(accuracy_info[module]["control_accuracy_55"] <= 0.5))

            ws.cell(16, ((module + 1) * 2 + 1), value=round(accuracy_info[module]["control_accuracy_72"], 2))
            ws.cell(16, ((module + 2) * 2), value=MainWindow.pf(accuracy_info[module]["control_accuracy_72"] <= 0.5))

            ws.cell(17, ((module + 1) * 2 + 1), value=round(accuracy_info[module]["control_accuracy_95"], 2))
            ws.cell(17, ((module + 2) * 2), value=MainWindow.pf(accuracy_info[module]["control_accuracy_95"] <= 0.5))

        for module in range(0, 4):
            ws.cell(18, ((module + 1) * 2 + 1),
                    value=accuracy_info[module]["temperature_accuracy_55"])
            ws.cell(18, ((module + 2) * 2),
                    value=MainWindow.pf(accuracy_info[module]["temperature_accuracy_55"]  <= 0.5))
            ws.cell(19, ((module + 1) * 2 + 1),
                    value=accuracy_info[module]["temperature_accuracy_72"])
            ws.cell(19, ((module + 2) * 2),
                    value=MainWindow.pf(accuracy_info[module]["temperature_accuracy_72"]  <= 0.5))
            ws.cell(20, ((module + 1) * 2 + 1),
                    value=accuracy_info[module]["temperature_accuracy_95"])
            ws.cell(20, ((module + 2) * 2),
                    value=MainWindow.pf(accuracy_info[module]["temperature_accuracy_95"] <= 0.5))

        for module in range(0, 4):
            ws.cell(21, ((module + 1) * 2 + 1), value=f'{round(time_accuracy_info[module]["time_accuracy"], 2)}%')
            ws.cell(21, ((module + 2) * 2),
                    value=MainWindow.pf(round(time_accuracy_info[module]["time_accuracy"], 2) <= 5))

        sheet_names = wb.get_sheet_names()
        for item in sheet_names:
            if item not in sheet_name_list:
                remove_sheet = wb.get_sheet_by_name(item)
                wb.remove_sheet(remove_sheet)

        wb.save(export_path)
        return True, "报告生成成功"



if __name__ == '__main__':
    multiprocessing.freeze_support()
    app = QtWidgets.QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec_())
    # 温度数据


